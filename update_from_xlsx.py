import openpyxl as ope
from tqdm import tqdm

TAB='\t'
def strproc(a:str):
    return a.replace('\t',' ').replace('\n',' ')

def load_xlsx():
    xlsx_path = "src.xlsx"
    workbook = ope.load_workbook(xlsx_path, data_only=True)
    worksheet = workbook.active
    r_index = 1
    res = []
    def read_cell(r_index, i):
        return strproc(worksheet.cell(r_index, i).value) if worksheet.cell(r_index, i).value else ''
    while worksheet.cell(r_index, 3).value:
        res.append([
            read_cell(r_index, 1),
            read_cell(r_index, 3),
            read_cell(r_index, 4)
        ])
        r_index += 1
    return res

def search(xlsx_data, ls):
    for line in xlsx_data:
        if line[0] == ls[0] and line[1] == ls[2]:
            return line[2]

if __name__ == "__main__":
    data={}
    origin_lines = open("SimplifiedChinese.tsv", encoding="utf-8").readlines()
    reference = load_xlsx()
    for index, line in tqdm(enumerate(origin_lines)):
        line = line.replace('\n','')
        ls = line.split('\t')
        if index > 0 and len(ls[3]) == 0 and len(ls[2]) != 0:
            origin_lines[index] = line + search(reference, ls) + '\n'

    with open("res.tsv", "w", encoding="utf-8") as f:
        for line in origin_lines:
            print(line, file=f, end='')
